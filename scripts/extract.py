"""
scripts/extract.py

Reads the source Excel workbook and writes data/contracts.json.

The set of data products to extract is defined in the DATA_PRODUCTS list below.
To add or remove a contract, edit that list and the corresponding sheet in the workbook.

Usage:
    python scripts/extract.py
    python scripts/extract.py --source path/to/file.xlsb --output data/contracts.json

Requires: LibreOffice (headless) for .xlsb -> .xlsx conversion + openpyxl.
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

# --------------------------------------------------------------------------- #
# Curated data product manifest. Edit this list to add/remove contracts.      #
# --------------------------------------------------------------------------- #
# Each tuple: (data_product, lives, table_name, sheet_name)
DATA_PRODUCTS: list[tuple[str, str, str, str]] = [
    ('APIMS Process', 'IRIS', 'product_xref', '30 Product Xref'),
    ('HCP Call Plan', 'IRIS', 'terr_hcp_call_plan', 'Terr_HCP_Call_Plan'),
    ('CET Tactic Metadata', 'IRIS', 'ibu_tactic_metadata', '97 CAP Campaign Metadata toEDB '),
    ('EDB Content XREF', 'IRIS', 'edb_content_xref', 'EDB CONTENT XREF'),
    ('HCP Tier', 'SFMC DE', 'Terr_HCP_Tier', 'HCP Tier'),
    ('HCP Rep Alignment', 'IRIS', 'hcp_rep_alignment', 'HCP Terr 2.0'),
    ('Hierarchy 1.0', 'IRIS', 'master_hierarchy_terr', '38 EIM_Master_Hierarchy_Terr'),
    ('SOA Brand Emails', 'IRIS', 'soa_brand_emails', '1 SOA_Brand-Sent Emails'),
    ('Lilly Events', 'IRIS', 'lilly_event_attendee', '109 Lilly Events'),
    ('Lilly Play Engagement', 'IRIS', 'lilly_play_engagement', 'lilly play engagement'),
    ('Lilly Play Activity', 'IRIS', 'lilly_play_activity', 'Lilly Play 3.0 Activity'),
    ('IBU Recommendations', 'S3 Refined Bucket',
     's3://lly-edp-refined-us-east-2-prod/iris/verso/out/verso_recommendations/Japan/',
     '91 VERSO NBE Reco(4C) Medical'),
    ('Verso Recommendation', 'S3 Refined Bucket',
     's3://lly-edp-refined-us-east-2-prod/iris/verso/out/doximity/\n'
     's3://lly-edp-refined-us-east-2-prod/iris/verso/out/medscape/\n'
     's3://lly-edp-refined-us-east-2-prod/iris/verso/out/epocrates/recommendations/',
     '49 NBE Recommendations(4C)'),
    ('Scout Process', 'IRIS', 'scout_global_restricted', '85 Scout_Standing_Request'),
    ('HCP Segmentation', 'IRIS', 'hcp_segmentation', 'HCP_Segmentation'),
    ('Sent Email Vod C', 'SFMC DE', 'Sent_Email_Vod__C', 'sent_email'),
    ('Universal HCO', 'IRIS', 'hco_universe', 'account'),
    ('Call2 Vod C', 'SFMC DE', 'Call_Vod_C', 'Call2_vod___c'),
    ('Lilly Play Content', 'IRIS', 'lilly_play_content', 'Lilly Play 3.0 Content'),
    ('Universal Customer List', 'IRIS', 'hcp_universe', '84 Universal Customer List'),
    ('Customer Data Processing', 'SFMC DE', 'Master_Customer_Data\nMaster_Profile_Data', '45 HCP Profile Data'),
    ('GCCP Consent', 'IRIS', 'lilly_program_opts', '83 GCCP- Consents'),
    ('Fam Formulary Access', 'IRIS', 'fam_formulary_access', 'Formulary Access'),
    ('Fam Formulary Brand', 'IRIS', 'fam_formulary_brand_indications', 'Formulary - Brand and Indicatio'),
    ('Fam Formulary Positioning', 'IRIS', 'fam_formulary_positioning', 'Formulary Positioning'),
    # LATAM
    ('LATAM App Downloads', 'IRIS', 'lilly_iris_zaidyn_prod.app_downloads', 'Latam - app_downloads'),
    ('LATAM Closeup Prescriptions', 'IRIS', 'lilly_iris_zaidyn_prod.d_closeup_prescriptions', 'Latam - d_closeup_prescriptions'),
    ('LATAM Discount Program Activity', 'IRIS', 'lilly_iris_zaidyn_prod.discount_program_activity', 'Latam - discount_program_activi'),
    ('LATAM Lilly360', 'IRIS', 'lilly_iris_zaidyn_prod.lilly360', 'Latam - lilly360'),
    ('LATAM WhatsApp', 'IRIS', 'lilly_iris_zaidyn_prod.whatsapp', 'Latam - whatsapp'),
    ('LATAM Lilly Online Store', 'IRIS', 'lilly_iris_zaidyn_prod.lilly_online_store', 'Latam - lilly_online_store'),
    ('LATAM SOA', 'IRIS', 'lilly_iris_zaidyn_prod.soa', 'Latam - soa'),
]

# --------------------------------------------------------------------------- #
# Extraction primitives                                                        #
# --------------------------------------------------------------------------- #

META_KEYS = {
    'File Name', 'Object Name', 'Description', 'Frequency', 'Delimiter', 'Location',
    'Full Feed/Delta', 'Product Owner', 'Responsible Delivery Team',
    'Feed Failure Notifications', 'Contract Last Updated', 'STM Link',
    'Redshift Table Name', 'Notes', 'NOTE', 'Last Verified/Updated',
    'Source System', 'Grain', 'Responsible Team/Platform',
}
SCHEMA_HEADERS = {'Column Name', 'col_name', 'Column Name '}


def clean_val(v) -> str:
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip().replace('\xa0', ' ').replace('\u00a0', ' ')


def find_schema_header(rows):
    for i, r in enumerate(rows):
        for j, c in enumerate(r):
            if c is not None and isinstance(c, str) and c.strip() in SCHEMA_HEADERS:
                return i, j
    return None, None


def detect_veeva_format(rows):
    """For sheets like sent_email/account that use a region-flag matrix."""
    REGION_KEYS = ['CMRCL', 'MEDICAL', 'IBU', ' DT', 'EU ', 'US ', 'CAN ', 'APAC']
    for i, r in enumerate(rows):
        if not r:
            continue
        first_idx = next((j for j, c in enumerate(r) if c is not None), None)
        if first_idx is None:
            continue
        region_cells = []
        for j in range(first_idx + 1, len(r)):
            c = r[j]
            if c is None:
                continue
            if any(k in str(c).upper() for k in REGION_KEYS):
                region_cells.append(str(c).strip())
        if len(region_cells) >= 3:
            return i, first_idx, [str(r[first_idx]).strip()] + region_cells
    return None, None, None


def extract_meta(rows, max_rows: int = 15) -> dict:
    meta: dict = {}
    for r in rows[:max_rows]:
        for i in range(len(r)):
            cell = r[i]
            if cell is None or not isinstance(cell, str):
                continue
            key = cell.strip()
            if key in META_KEYS and key not in meta:
                for j in range(i + 1, len(r)):
                    if r[j] is not None:
                        meta[key] = clean_val(r[j])
                        break
    return meta


def extract_contract(wb, sheet_name: str):
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    meta = extract_meta(rows)
    schema_headers: list[str] = []
    columns: list[dict] = []

    hdr_row, hdr_col = find_schema_header(rows)
    if hdr_row is not None:
        header = list(rows[hdr_row][hdr_col:])
        clean_hdr = [clean_val(h) if h is not None else '' for h in header]
        while clean_hdr and not clean_hdr[-1]:
            clean_hdr.pop()
        clean_hdr = [h or f'_col{i}' for i, h in enumerate(clean_hdr)]
        clean_hdr[0] = 'Column Name'
        schema_headers = clean_hdr
        for r in rows[hdr_row + 1:]:
            if hdr_col >= len(r):
                continue
            data = r[hdr_col:hdr_col + len(clean_hdr)]
            if not any(x is not None for x in data):
                continue
            first = data[0]
            if first is None:
                continue
            if isinstance(first, str):
                fs = first.strip()
                if not fs or fs in META_KEYS or fs in SCHEMA_HEADERS:
                    continue
            col = {}
            for i, key in enumerate(clean_hdr):
                if i < len(data) and data[i] is not None:
                    col[key] = clean_val(data[i])
            if col.get('Column Name'):
                columns.append(col)
    else:
        v_row, v_col, regions = detect_veeva_format(rows)
        if v_row is not None:
            region_cols = regions[1:]
            schema_headers = ['Column Name'] + region_cols
            for r in rows[v_row + 1:]:
                if v_col >= len(r):
                    continue
                first = r[v_col]
                if first is None:
                    continue
                fs = clean_val(first)
                if not fs or fs in META_KEYS:
                    continue
                col = {'Column Name': fs}
                for i, key in enumerate(region_cols, start=1):
                    idx = v_col + i
                    col[key] = clean_val(r[idx]) if idx < len(r) and r[idx] is not None else ''
                columns.append(col)

    return meta, schema_headers, columns


def feed_type(dp: str) -> str:
    n = dp.lower()
    if any(k in n for k in ('xref', 'apims', 'hierarchy', 'tier')):
        return 'MASTER'
    if 'metadata' in n or 'content' in n:
        return 'METADATA'
    if 'consent' in n:
        return 'CONSENT'
    return 'DATA FEED'


def workstream(dp: str) -> str:
    n = dp.lower()
    if 'latam' in n: return 'LATAM'
    if 'formulary' in n: return 'PAYER'
    if 'lilly play' in n: return 'LILLY PLAY'
    if 'verso' in n or 'recommendation' in n: return 'VERSO'
    if 'soa' in n or 'email' in n: return 'HQ EMAIL & SOA'
    if 'scout' in n: return 'MEDICAL'
    return 'OMNICHANNEL'


def source_system(dp: str) -> str:
    n = dp.lower()
    if 'latam' in n: return 'LATAM Region'
    if 'apims' in n: return 'APIMS'
    if 'lilly play' in n: return 'Lilly Play'
    if 'cap' in n or 'tactic' in n: return 'CAP'
    if 'edb' in n: return 'EDB'
    if any(k in n for k in ('hcp tier', 'hierarchy', 'rep alignment', 'segmentation')): return 'EIM'
    if 'verso' in n or 'recommendation' in n: return 'Verso'
    if 'scout' in n: return 'Scout'
    if any(k in n for k in ('sent email', 'call2', 'universal hco', 'customer data')): return 'Veeva CRM'
    if 'event' in n: return 'Veeva Events'
    if 'soa' in n: return 'SOA Engine'
    if 'gccp' in n or 'consent' in n: return 'GCCP'
    if 'formulary' in n: return 'Fam (Formulary)'
    return 'EDB'


def norm_load(v: str) -> str:
    if not v:
        return 'Unknown'
    s = v.strip().lower()
    if 'full' in s: return 'Full'
    if 'incremental' in s: return 'Incremental'
    if 'delta' in s: return 'Delta'
    return v.strip()


def convert_xlsb_to_xlsx(xlsb_path: Path) -> Path:
    """Use LibreOffice headless to convert .xlsb to .xlsx."""
    out_dir = Path(tempfile.mkdtemp())
    cmd = ['libreoffice', '--headless', '--convert-to', 'xlsx',
           '--outdir', str(out_dir), str(xlsb_path)]
    subprocess.run(cmd, check=True, capture_output=True, timeout=180)
    return out_dir / (xlsb_path.stem + '.xlsx')


def build_contracts(workbook_path: Path) -> list[dict]:
    if workbook_path.suffix.lower() == '.xlsb':
        workbook_path = convert_xlsb_to_xlsx(workbook_path)
    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    contracts = []
    for idx, (dp, lives, tn, sheet) in enumerate(DATA_PRODUCTS, start=1):
        if sheet not in wb.sheetnames:
            print(f"WARNING: sheet '{sheet}' for '{dp}' not found in workbook", file=sys.stderr)
            continue
        meta, hdrs, cols = extract_contract(wb, sheet)
        contracts.append({
            'sheet': sheet,
            'displayName': dp,
            'numPrefix': idx,
            'fileName': clean_val(meta.get('File Name', '') or meta.get('Object Name', '')),
            'description': clean_val(meta.get('Description', '')),
            'frequency': clean_val(meta.get('Frequency', '')),
            'delimiter': clean_val(meta.get('Delimiter', '')),
            'location': clean_val(meta.get('Location', '')),
            'loadType': norm_load(meta.get('Full Feed/Delta', '')),
            'notes': clean_val(meta.get('Notes', '') or meta.get('NOTE', '')),
            'productOwner': clean_val(meta.get('Product Owner', '')),
            'deliveryTeam': clean_val(meta.get('Responsible Delivery Team', '') or meta.get('Responsible Team/Platform', '')),
            'redshiftTable': clean_val(meta.get('Redshift Table Name', '')),
            'lastVerified': clean_val(meta.get('Contract Last Updated', '') or meta.get('Last Verified/Updated', '')),
            'columnCount': len(cols),
            'feedType': feed_type(dp),
            'workstream': workstream(dp),
            'sourceSystem': source_system(dp),
            'status': 'ACTIVE',
            'schemaHeaders': hdrs,
            'columns': cols,
            'dataProduct': dp,
            'lives': lives,
            'tableName': tn,
        })
    return contracts


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--source', default='source/Current_State_Data_Inventory.xlsb',
                        help='Path to the .xlsb or .xlsx source workbook')
    parser.add_argument('--output', default='data/contracts.json',
                        help='Path to write the contracts JSON')
    args = parser.parse_args()

    src = Path(args.source)
    out = Path(args.output)

    if not src.exists():
        print(f"ERROR: source workbook not found: {src}", file=sys.stderr)
        return 1

    contracts = build_contracts(src)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(contracts, indent=2, default=str), encoding='utf-8')

    total_cols = sum(c['columnCount'] for c in contracts)
    print(f"✓ Wrote {len(contracts)} contracts ({total_cols} schema columns) to {out}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
